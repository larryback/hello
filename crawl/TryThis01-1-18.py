#    from pyexcel.cookbook import merge_all_to_a_book
    # import pyexcel.ext.xlsx # no longer required if you use pyexcel >= 0.2.2 
#    import glob


#    merge_all_to_a_book(glob.glob("your_csv_directory/*.csv"), "output.xlsx")

#import pandas as pd

#filepath_in = "./melon_top_100.csv"
#filepath_out = "./Melon_top_100.xlsx"
#pd.read_csv(filepath_in, delimiter=";", encoding = euc-kr).to_excel(filepath_out)

#pd.read_csv('./melon_top_100.csv')
#.to_excel('Melon_top_100.xlsx')

    # import openpyxl 
    # import csv 

    # examplefile = open('./melon_top_100.csv') 
    # exampleReader = csv.reader(examplefile) 
    # exampleData = list(exampleReader) 


    # book = openpyxl.load_workbook('./Melon_top_100.xlsx') 
    # sheet = book.worksheets[0]

    # for i in range (1,9): 
    #     for h in range(1, 5): 
    #     a = i-1 
    #     b = h-1 
    #     ws.cell(row=i, column=h).value = exampleData[a][b] 

# CSV 읽기
import openpyxl
import csv, codecs

#wb.save('./melon_top_100.csv') 
import openpyxl
import csv, codecs

fp = codecs.open("./melon_top_100.csv", "r", "utf-8")

# aaa,bbb,"ccc,cc"
reader = csv.reader(fp, delimiter=',', quotechar='"')

print("-------------------------------------------------------------")
print(reader)

#  XLSX 쓰기
import openpyxl
book = openpyxl.Workbook(reader)
sheet1 = book.active
    # sheet1.title = "첫번째 시트"
    # sheet1.cell(row=1, column=1).value = 'Title'
    # sheet2 = book.create_sheet()
    # sheet2.title = "두번째 시트"
    # sheet2['A1'] = datetime.datetime.now()
    # sheet2['A2'] = datetime.date.today()
# 저장하기
book.save("./Melon_top_100.xlsx")

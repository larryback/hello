import csv, codecs
import random
import openpyxl
from openpyxl import Workbook
<<<<<<< HEAD
from openpyxl.cell.cell import get_column_letter
=======
import time
>>>>>>> 14b1a64526eeb51aecda34d9ad38c7642b2542db

fp = codecs.open("./melon_top_100.csv", "r", "ms949")

# aaa,bbb,"ccc,cc"
reader = csv.reader(fp, delimiter=',', quotechar='"')

<<<<<<< HEAD
print(list(reader))
=======
# print(list(reader))
>>>>>>> 14b1a64526eeb51aecda34d9ad38c7642b2542db


#book = openpyxl.Workbook()
#sheet1 = book.active
#sheet1.title = "첫번째 시트"


book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "첫번째 시트"


<<<<<<< HEAD
wb = Workbook()
ws = wb.worksheets[0]
ws.title = "첫번째 페이지"

for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

wb.save("./melontop100.xlsx")






#with codecs.open("./melon_top_100.csv", "r", "CP949" ) as f:
    # book = openpyxl.Workbook()
    # reader = csv.reader(fp, delimiter=',', quotechar='"')
    # table = reader = csv.reader(fp, delimiter=',', quotechar='"')
    # for row in table:

    # ws.write_row(i, 0, row)
    # i += 1
    # wb.close()

#for r, row in enumerate(reader):
#    for c, col in enumerate(row):


#for idx, val in enumerate(col.split(CSV_SEPARATOR)):
#cell = sheet1.cell(row=r+1, column=idx+1)
#cell.value = val

    # book.save("./melontop100.xlsx")

    # convert_csv_to_xlsx() 


# this may be adjusted to use 'excel types' explicitly (see xlsxwriter doc)
=======


for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):

       print(cell) 
       sheet1.cell(row = row + 1, column=1).value = '랭킹'
       #sheet1.cell(row=1, column=1).value = ((column_index + 1))
       #sheet1.cell(row=1, column=1).value = ((row_index + 1))

book.save("./melontop100.xlsx")



>>>>>>> 14b1a64526eeb51aecda34d9ad38c7642b2542db





<<<<<<< HEAD
# with codecs.open('./melontop100.xlsx', 'w', 'CP949') as ff:
#     writer = csv.writer(ff, delimiter=',', quotechar='"')

#     for cells in reader:
#         writer.writerow([cells[0], random.randrange(1,100)])    
=======

>>>>>>> 14b1a64526eeb51aecda34d9ad38c7642b2542db

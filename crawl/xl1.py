import openpyxl
import datetime

book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "첫번째 시트"
sheet1.cell(row=1, column=1).value = 'Title'
sheet1.cell(row=4, column=4).value = '백영래'
sheet2 = book.create_sheet()
sheet2.title = "두번째 시트"
sheet2['A1'] = datetime.datetime.now()
sheet2['A2'] = datetime.date.today()
# 저장하기
#book.save("./output01.xlsx")

import openpyxl
# insert image
imgFile = './crawl/images/aaa.png'
img = openpyxl.drawing.image.Image(imgFile)
sheet1.add_image(img, 'B5')
# resize image
from PIL import Image
img2 = Image.open(imgFile)
new_img = img2.resize((50, 50))
new_img.save('new.png')
img3 = openpyxl.drawing.image.Image('new.png')
sheet1.add_image(img3, 'B5')
book.save("./output01.xlsx")
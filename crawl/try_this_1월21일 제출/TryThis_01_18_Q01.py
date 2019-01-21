import csv, codecs
import random
import openpyxl
from openpyxl import Workbook
import time

fp = codecs.open("./melon_top_100.csv", "r", "ms949")

# aaa,bbb,"ccc,cc"
reader = csv.reader(fp, delimiter=',', quotechar='"')

# print(list(reader))


#book = openpyxl.Workbook()
#sheet1 = book.active
#sheet1.title = "첫번째 시트"


book = openpyxl.Workbook()
sheet1 = book.active
sheet1.title = "첫번째 시트"




for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):

       print(cell) 
      
       #sheet1.cell(row=1, column=1).value = ((column_index + 1))
       #sheet1.cell(row=1, column=1).value = ((row_index + 1))

book.save("./melontop100.xlsx")









